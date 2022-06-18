import requests
import json
import jsonpath
import openpyxl

class Common:
   
  def __init__(self):
    global vk
    global sh
    vk = openpyxl.load_workbook(FileNamePath)
    sh = vk[SheetName]
      
  def fetch_row_count(self, FileNamePath, SheetName):
    max_rows = sh.max_row
    return max_rows
  
  
  def fetch_column_count(self, FileNamePath,SheetName):
    max_columns = sh.max_column
    return max_columns
   
   def update_request_with_data(self, rowNumber, jsonRequest, keyList):
      c = sh.max_column
      for i in range(1, c+1):
         cell = sh.cell(row = rowNumber, column = i)
         jsonRequest[keyList[i-1]] = cell.value
      return jsonRequest
   
   
         
