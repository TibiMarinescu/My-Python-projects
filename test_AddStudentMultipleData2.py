import requests
import json
import jsonpath
import openpyxl

def test_add_multiple_students():
   f = open("C:/Users/tibi/Documents/Training/REST API/JSON_input.json",'r')
   json_input = json.loads(f.read())
   vk = openpyxl.load_workbook("C:/Users/tibi/Documents/Training/REST API/Input_file.xlsx")
   sh = vk['Sheet1']
   rows = sh.max_row
   for i in range(2, rows+1):
        cell_first_name = sh.cell(row=i,column=1)
        cell_middle_name = sh.cell(row=i,column=2)
        cell_last_name = sh.cell(row=i,column=3)
        cell_date_of_birth = sh.cell(row=i,column=4)
        json_input['first_name'] = cell_first_name.value
        json_input['middle_name'] = cell_middle_name.value
        json_input['last_name'] = cell_last_name.value
        json_input['date_of_birth'] = cell_date_of_birth.value
